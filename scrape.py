"""
OnlineJobs.ph Keyword Job Scraper
=================================
Scrapes public job listings from www.onlinejobs.ph for given keywords
and exports an Excel file with:

  - Job Post Title
  - Salary
  - Link
  - Skill Requirements
  - Employer Info
  - Date Updated  (from each job post)
  - Posted Date, Employment Type, Matched Keyword

When you run `python scrape.py` with no flags, it interactively asks for:
  1) keywords to search
  2) max results (e.g. 50 = latest 50 after sort)

Personal / research use only. Respect robots.txt crawl-delay and site ToS.

Usage:
    python scrape.py
    python scrape.py --keywords "virtual assistant,VA" --max-results 50
    python scrape.py --max-results 30 --delay 3
"""

from __future__ import annotations

import argparse
import math
import re
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Iterable
from urllib.parse import quote_plus, urljoin

import requests
from bs4 import BeautifulSoup

try:
    import pandas as pd
except ImportError:
    print("ERROR: pandas is required. Run: pip install -r requirements.txt")
    sys.exit(1)


BASE_URL = "https://www.onlinejobs.ph"
SEARCH_PATH = "/jobseekers/jobsearch"
OUTPUT_DIR = Path(__file__).resolve().parent / "output"

DEFAULT_KEYWORDS = [
    "virtual assistant",
    "VA",
    "Data Entry",
    "Customer Service",
]
DEFAULT_MAX_RESULTS = 50

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/120.0.0.0 Safari/537.36"
)

# robots.txt recommends Crawl-delay: 5
DEFAULT_DELAY = 5.0
PAGE_SIZE = 30

# Excel columns (order matters)
EXCEL_COLUMNS = [
    "Job Post Title",
    "Salary",
    "Link",
    "Skill Requirements",
    "Employer Info",
    "Date Updated",
    "Posted Date",
    "Employment Type",
    "Matched Keyword",
]


# ──────────────────────────────────────────────
# HTTP
# ──────────────────────────────────────────────

def make_session() -> requests.Session:
    session = requests.Session()
    session.headers.update(
        {
            "User-Agent": USER_AGENT,
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "en-US,en;q=0.9",
            "Referer": f"{BASE_URL}/",
        }
    )
    return session


def fetch(session: requests.Session, url: str, delay: float) -> str | None:
    """GET a URL with polite delay. Returns HTML or None on failure."""
    time.sleep(delay)
    try:
        resp = session.get(url, timeout=30)
        resp.raise_for_status()
        resp.encoding = resp.apparent_encoding or "utf-8"
        return resp.text
    except requests.RequestException as exc:
        print(f"  [warn] Failed to fetch {url}: {exc}")
        return None


# ──────────────────────────────────────────────
# Parsing
# ──────────────────────────────────────────────

def clean_text(value: str | None) -> str:
    if not value:
        return ""
    text = BeautifulSoup(value, "lxml").get_text(" ", strip=True)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def parse_total_jobs(html: str) -> int | None:
    m = re.search(r"Displaying\s+\d+\s+out of\s+(\d+)\s+jobs", html, re.I)
    if m:
        return int(m.group(1))
    return None


def parse_listing_page(html: str, keyword: str) -> list[dict]:
    """Parse job cards from a search results page."""
    soup = BeautifulSoup(html, "lxml")
    jobs: list[dict] = []

    for card in soup.select("a[href*='/jobseekers/job/'] > div.jobpost-cat-box"):
        parent = card.find_parent("a")
        if not parent:
            continue

        href = parent.get("href", "").strip()
        if not href or "/jobseekers/job/" not in href:
            continue

        link = urljoin(BASE_URL, href)
        job_id_match = re.search(r"-(\d+)/?$", href)
        job_id = job_id_match.group(1) if job_id_match else href

        h4 = card.select_one("h4")
        title = ""
        employment_type = ""
        if h4:
            badge = h4.select_one("span.badge")
            if badge:
                employment_type = clean_text(badge.get_text())
                badge.extract()
            title = clean_text(h4.get_text())

        salary = ""
        salary_icon = card.select_one("i.icon-round-dollar")
        if salary_icon:
            dt = salary_icon.find_parent("dt")
            if dt:
                sibling = dt.find_next_sibling("dd")
                if sibling:
                    salary = clean_text(sibling.get_text())
        if not salary:
            for dd in card.select("dd.col"):
                text = clean_text(dd.get_text())
                if text:
                    salary = text
                    break

        posted = ""
        # Prefer data-temp attribute (ISO-like datetime) when present
        p_el = card.select_one("p[data-temp]")
        if p_el and p_el.get("data-temp"):
            posted = p_el.get("data-temp", "").strip()
        else:
            em = card.select_one("p em")
            if em:
                posted = clean_text(em.get_text()).replace("Posted on ", "")

        skills = [
            clean_text(a.get_text())
            for a in card.select("div.job-tag a.badge")
            if clean_text(a.get_text())
        ]

        employer = ""
        logo = card.select_one("img.jobpost-cat-box-logo[alt]")
        if logo:
            alt = (logo.get("alt") or "").strip()
            if alt and alt.lower() not in {"oj logo here", "logo"}:
                employer = alt

        jobs.append(
            {
                "job_id": job_id,
                "Job Post Title": title,
                "Salary": salary,
                "Link": link,
                "Skill Requirements": "; ".join(skills),
                "Employer Info": employer,
                "Date Updated": "",  # filled from detail page
                "Posted Date": posted,
                "Employment Type": employment_type,
                "Matched Keyword": keyword,
            }
        )

    return jobs


def parse_date_updated(soup: BeautifulSoup) -> str:
    """Extract DATE UPDATED from a job detail page."""
    # Structure: <h3>DATE UPDATED</h3><p> Jul 23, 2026 </p>
    for h3 in soup.find_all(["h3", "h4", "strong", "label"]):
        label = clean_text(h3.get_text())
        if re.search(r"DATE\s*UPDATED", label, re.I):
            # Next sibling <p> or next text block
            p = h3.find_next("p")
            if p:
                return clean_text(p.get_text())
            parent = h3.find_parent()
            if parent:
                text = clean_text(parent.get_text())
                text = re.sub(r"(?i)DATE\s*UPDATED", "", text).strip()
                if text:
                    return text

    # Fallback regex on whole page text
    m = re.search(
        r"DATE\s*UPDATED\s*([A-Za-z]{3,9}\s+\d{1,2},?\s+\d{4}|\d{4}-\d{2}-\d{2})",
        soup.get_text(" ", strip=True),
        re.I,
    )
    if m:
        return m.group(1).strip()
    return ""


def enrich_from_detail(session: requests.Session, job: dict, delay: float) -> dict:
    """Open job detail page for Date Updated (+ skills/employer/salary if missing)."""
    html = fetch(session, job["Link"], delay)
    if not html:
        return job

    soup = BeautifulSoup(html, "lxml")

    date_updated = parse_date_updated(soup)
    if date_updated:
        job["Date Updated"] = date_updated

    skill_links = soup.select("a.card-worker-topskill")
    if skill_links:
        skills = [clean_text(a.get_text()) for a in skill_links if clean_text(a.get_text())]
        if skills:
            job["Skill Requirements"] = "; ".join(skills)

    if not job.get("Employer Info"):
        logo = soup.select_one("img[alt][src*='employer']")
        if logo:
            alt = (logo.get("alt") or "").strip()
            if alt:
                job["Employer Info"] = alt

        for heading in soup.find_all(string=re.compile(r"VIEW OTHER JOB POSTS FROM", re.I)):
            parent = heading.find_parent()
            if parent:
                nxt = parent.find_next(["a", "strong", "h5", "h4", "span"])
                if nxt:
                    name = clean_text(nxt.get_text())
                    if name and "share" not in name.lower():
                        job["Employer Info"] = name
                        break

    if not job.get("Salary"):
        for label in soup.find_all(string=re.compile(r"WAGE\s*/\s*SALARY", re.I)):
            block = label.find_parent()
            if block:
                parent = block.find_parent()
                if parent:
                    text = clean_text(parent.get_text())
                    text = re.sub(r"(?i)TYPE OF WORK.*", "", text)
                    text = re.sub(r"(?i)WAGE\s*/\s*SALARY", "", text).strip()
                    if text:
                        job["Salary"] = text.split("HOURS")[0].strip()
                        break

    return job


# ──────────────────────────────────────────────
# Scrape orchestration
# ──────────────────────────────────────────────

def search_url(keyword: str | None = None, offset: int = 0) -> str:
    """Build job search URL. Empty/None keyword = all jobs (no filter)."""
    kw = (keyword or "").strip()
    if offset <= 0:
        if not kw:
            return f"{BASE_URL}{SEARCH_PATH}"
        return f"{BASE_URL}{SEARCH_PATH}?jobkeyword={quote_plus(kw)}"
    if not kw:
        return f"{BASE_URL}{SEARCH_PATH}/{offset}"
    return f"{BASE_URL}{SEARCH_PATH}/{offset}?jobkeyword={quote_plus(kw)}"


def scrape_keyword(
    session: requests.Session,
    keyword: str | None,
    max_pages: int | None,
    delay: float,
    stop_after: int | None = None,
) -> list[dict]:
    """
    Scrape listing pages for one keyword (newest first on the site).
    Empty/None keyword = all latest jobs (unfiltered jobsearch).
    stop_after: stop collecting once this many jobs are gathered for this keyword.
    """
    label = (keyword or "").strip() or "(all jobs)"
    print(f"\n=== Keyword: {label!r} ===")
    all_jobs: list[dict] = []
    match_label = (keyword or "").strip() or "all"

    first_url = search_url(keyword, 0)
    print(f"  Page 1: {first_url}")
    html = fetch(session, first_url, delay)
    if not html:
        return all_jobs

    total = parse_total_jobs(html)
    if total is not None:
        print(f"  Site reports ~{total} jobs for this search")
    else:
        print("  Could not read total job count; will stop when a page is empty")

    page_jobs = parse_listing_page(html, match_label)
    print(f"  Found {len(page_jobs)} jobs on page 1")
    all_jobs.extend(page_jobs)

    if not page_jobs:
        return all_jobs

    total_pages = 1
    if total is not None:
        total_pages = max(1, (total + PAGE_SIZE - 1) // PAGE_SIZE)
    if max_pages is not None:
        total_pages = min(total_pages, max_pages)
    if stop_after is not None:
        needed_pages = max(1, math.ceil(stop_after / PAGE_SIZE))
        total_pages = min(total_pages, needed_pages)

    for page_num in range(2, total_pages + 1):
        if stop_after is not None and len(all_jobs) >= stop_after:
            break
        offset = (page_num - 1) * PAGE_SIZE
        url = search_url(keyword, offset)
        print(f"  Page {page_num}/{total_pages}: offset={offset}")
        html = fetch(session, url, delay)
        if not html:
            break
        page_jobs = parse_listing_page(html, match_label)
        print(f"  Found {len(page_jobs)} jobs")
        if not page_jobs:
            break
        all_jobs.extend(page_jobs)

    if stop_after is not None:
        all_jobs = all_jobs[:stop_after]
    return all_jobs


def dedupe_jobs(jobs: Iterable[dict]) -> list[dict]:
    """Keep first occurrence per job_id; merge matched keywords."""
    by_id: dict[str, dict] = {}
    for job in jobs:
        jid = job.get("job_id") or job.get("Link")
        if jid not in by_id:
            by_id[jid] = dict(job)
            continue
        existing = by_id[jid]
        keys = {
            k.strip()
            for k in (
                existing.get("Matched Keyword", "") + "," + job.get("Matched Keyword", "")
            ).split(",")
            if k.strip()
        }
        existing["Matched Keyword"] = ", ".join(sorted(keys, key=str.lower))
        for field in (
            "Salary",
            "Skill Requirements",
            "Employer Info",
            "Employment Type",
            "Posted Date",
            "Date Updated",
        ):
            if not existing.get(field) and job.get(field):
                existing[field] = job[field]
    return list(by_id.values())


def sort_key_datetime(job: dict) -> datetime:
    """Best available datetime for sorting (Date Updated preferred)."""
    for field in ("Date Updated", "Posted Date"):
        raw = (job.get(field) or "").strip()
        if not raw:
            continue
        # Try common formats
        for fmt in (
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d",
            "%b %d, %Y",
            "%B %d, %Y",
            "%b %d %Y",
            "%B %d %Y",
        ):
            try:
                return datetime.strptime(raw, fmt)
            except ValueError:
                continue
        parsed = pd.to_datetime(raw, errors="coerce")
        if pd.notna(parsed):
            return parsed.to_pydatetime()
    return datetime.min


def sort_jobs_latest_first(jobs: list[dict]) -> list[dict]:
    return sorted(jobs, key=sort_key_datetime, reverse=True)


def export_excel(jobs: list[dict], path: Path) -> Path:
    rows = [{col: job.get(col, "") or "" for col in EXCEL_COLUMNS} for job in jobs]
    df = pd.DataFrame(rows, columns=EXCEL_COLUMNS).fillna("")

    # Sort newest first: Date Updated, then Posted Date
    df["_sort"] = pd.to_datetime(df["Date Updated"], errors="coerce")
    missing = df["_sort"].isna()
    if missing.any():
        df.loc[missing, "_sort"] = pd.to_datetime(
            df.loc[missing, "Posted Date"], errors="coerce"
        )
    df = df.sort_values("_sort", ascending=False, na_position="last").drop(columns=["_sort"])

    path.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(path, index=False, engine="openpyxl")

    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter

        wb = load_workbook(path)
        ws = wb.active
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for idx, col_name in enumerate(EXCEL_COLUMNS, start=1):
            max_len = len(col_name)
            for cell in ws[get_column_letter(idx)]:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), 60))
            ws.column_dimensions[get_column_letter(idx)].width = max_len + 2
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"
        wb.save(path)
    except Exception as exc:
        print(f"  [warn] Could not style Excel file: {exc}")

    return path


# ──────────────────────────────────────────────
# Interactive prompts + CLI
# ──────────────────────────────────────────────

def prompt_keywords(default: list[str]) -> list[str]:
    default_str = ", ".join(default)
    print("\nEnter keywords to search (comma-separated).")
    print(f"  Press Enter for default: {default_str}")
    raw = input("Keywords: ").strip()
    if not raw:
        return list(default)
    return [k.strip() for k in raw.split(",") if k.strip()]


def prompt_max_results(default: int = DEFAULT_MAX_RESULTS) -> int:
    print("\nHow many latest results do you want in the Excel file?")
    print(f"  Example: 50  → newest 50 jobs after sorting by Date Updated")
    print(f"  Press Enter for default: {default}")
    while True:
        raw = input("Max results: ").strip()
        if not raw:
            return default
        try:
            n = int(raw)
            if n <= 0:
                print("  Please enter a positive number.")
                continue
            return n
        except ValueError:
            print("  Please enter a whole number (e.g. 50).")


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Scrape OnlineJobs.ph listings by keyword into Excel."
    )
    p.add_argument(
        "--keywords",
        default=None,
        help='Comma-separated keywords. If omitted, you will be prompted.',
    )
    p.add_argument(
        "--max-results",
        type=int,
        default=None,
        help=f"Keep only the latest N jobs (default when prompted: {DEFAULT_MAX_RESULTS}).",
    )
    p.add_argument(
        "--max-pages",
        type=int,
        default=None,
        help=f"Optional cap on listing pages per keyword ({PAGE_SIZE} jobs/page).",
    )
    p.add_argument(
        "--delay",
        type=float,
        default=DEFAULT_DELAY,
        help=f"Seconds between requests (default {DEFAULT_DELAY}).",
    )
    p.add_argument(
        "--skip-detail",
        action="store_true",
        help="Do not open job pages (faster, but Date Updated will be empty).",
    )
    p.add_argument(
        "--no-prompt",
        action="store_true",
        help="Do not ask interactively; use defaults / flags only.",
    )
    p.add_argument(
        "--output",
        default=None,
        help="Output .xlsx path (default: output/onlinejobs_YYYYMMDD_HHMMSS.xlsx)",
    )
    return p.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)

    # Interactive when flags are missing (unless --no-prompt)
    interactive = not args.no_prompt and (args.keywords is None or args.max_results is None)

    if interactive:
        print("=" * 56)
        print("  OnlineJobs.ph Job Scraper")
        print("=" * 56)
        print("Results go to an Excel file, sorted newest first.")
        print("Date Updated is read from each job post (detail page).")

    if args.keywords is not None:
        keywords = [k.strip() for k in args.keywords.split(",") if k.strip()]
    elif args.no_prompt:
        keywords = list(DEFAULT_KEYWORDS)
    else:
        keywords = prompt_keywords(DEFAULT_KEYWORDS)

    if not keywords:
        print("No keywords provided.")
        return 1

    if args.max_results is not None:
        max_results = args.max_results
        if max_results <= 0:
            print("--max-results must be a positive integer.")
            return 1
    elif args.no_prompt:
        max_results = DEFAULT_MAX_RESULTS
    else:
        max_results = prompt_max_results(DEFAULT_MAX_RESULTS)

    fetch_detail = not args.skip_detail

    print("\nStarting scrape…")
    print(f"  Keywords    : {keywords}")
    print(f"  Max results : {max_results} (latest after sort)")
    print(f"  Delay       : {args.delay}s between requests")
    print(f"  Detail pages: {'yes (for Date Updated)' if fetch_detail else 'no'}")
    print("  Note        : Personal job-hunting use only. Respect site ToS.")

    session = make_session()
    collected: list[dict] = []

    # Per keyword, pull enough listing pages to cover max_results
    # (a bit of buffer helps after multi-keyword dedupe)
    per_kw_target = max_results

    for kw in keywords:
        try:
            collected.extend(
                scrape_keyword(
                    session,
                    kw,
                    max_pages=args.max_pages,
                    delay=args.delay,
                    stop_after=per_kw_target,
                )
            )
        except KeyboardInterrupt:
            print("\nInterrupted during keyword scrape. Saving what we have…")
            break

    jobs = dedupe_jobs(collected)
    print(f"\nUnique jobs after listing + dedupe: {len(jobs)}")

    # Pre-sort by listing Posted Date so we only open the newest candidates
    jobs = sort_jobs_latest_first(jobs)
    candidates = jobs[:max_results] if max_results else jobs
    print(f"Keeping top {len(candidates)} candidates for detail enrichment / export")

    if fetch_detail and candidates:
        print(f"\nOpening {len(candidates)} job posts for Date Updated…")
        for i, job in enumerate(candidates, start=1):
            title = (job.get("Job Post Title") or "")[:55]
            print(f"  [{i}/{len(candidates)}] {title}")
            try:
                enrich_from_detail(session, job, delay=args.delay)
            except KeyboardInterrupt:
                print("\nInterrupted during detail scrape.")
                break

    # Final sort by Date Updated (fallback Posted Date), then cap
    candidates = sort_jobs_latest_first(candidates)
    if max_results:
        candidates = candidates[:max_results]

    if not candidates:
        print("No jobs found. Nothing to export.")
        return 2

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = Path(args.output) if args.output else OUTPUT_DIR / f"onlinejobs_{stamp}.xlsx"
    export_excel(candidates, out_path)

    with_employer = sum(1 for j in candidates if j.get("Employer Info"))
    with_skills = sum(1 for j in candidates if j.get("Skill Requirements"))
    with_updated = sum(1 for j in candidates if j.get("Date Updated"))

    print("\nDone.")
    print(f"  Excel file      : {out_path}")
    print(f"  Rows            : {len(candidates)} (latest first)")
    print(f"  With Date Updated: {with_updated}")
    print(f"  With skills     : {with_skills}")
    print(f"  With employer   : {with_employer}")
    print(
        "  Note: Employer name is only public when the listing shows a company logo."
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
